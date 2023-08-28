from openpyxl import load_workbook
input_file = 'C:\\Users\\29844\\Desktop\\福建省 region_code.xlsx'
output_file = 'C:\\Users\\29844\\Desktop\\福建省 region_code.xlsx'
sheet_name = 'Sheet1'  # 工作表名称

# 修改3，5，7列
column_index_to_modify = 7  # 要修改的列的索引（从1开始）
old_prefix = '350605'
new_prefix = '350625'

workbook = load_workbook(input_file)
sheet = workbook[sheet_name]

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=column_index_to_modify, max_col=column_index_to_modify):
    cell = row[0]
    if cell.value is not None and str(cell.value).startswith(old_prefix):
        cell.value = new_prefix + str(cell.value)[len(old_prefix):]

workbook.save(output_file)
print("XLSX column processing complete.")